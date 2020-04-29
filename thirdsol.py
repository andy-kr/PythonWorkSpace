import csv
from statistics import mean

import openpyxl
from openpyxl.chart import BarChart, Reference, LineChart

import sqlite3

data = []

def makeData():
    fp = open('CCTV_in_Seoul.csv', encoding='utf-8')
    rd = csv.reader(fp)
    next(rd)    # remove first line

    for r in rd:
        data.append((r[0], int(r[1]), int(r[2]), int(r[3]), int(r[4]), int(r[5])))
#print(data)
total = 0

makeData()

# 1
avg = mean([n[1] for n in data])
print("평균 = ", avg)

def totMax():
    mx = max(data, key=lambda v:v[1])
    print(mx[0], mx[1])

totMax()

# 2
# max_num=0
# max_num_gu = ''
# for n in data:
#     if max_num < n[1]:
#         max_num = n[1]
#         max_num_gu = n[0]
#
# print(max_num_gu, max_num)

#3 sorted
sorted_list = sorted(n[1] for n in data)
sorted_list.reverse()
for i in range(0,5):
    print(sorted_list[i])

def totTop5():
    sData = sorted(data, key=lambda v:v[1], reverse=True)[0:5]
    #print(s)
    for n in sData:
        print(n[0], n[1])

totTop5()

def excelChart():
    wb=openpyxl.Workbook()
    sh=wb.active

    sh.append(['기관명', '소계'])
    for dt in data:
        sh.append([dt[0], dt[1]])

    chart = LineChart()
    cdata = Reference(sh, min_col=2, min_row=1, max_row=sh.max_row)
    cat = Reference(sh, min_col=1, min_row=2, max_row=sh.max_row)

    chart.add_data(cdata, titles_from_data=True)
    chart.set_categories(cat)
    sh.add_chart(chart, 'F2')

    wb.save('cctv_sum.xlsx')

excelChart()

#4
# wb = openpyxl.Workbook()
# sh = wb.active
#
# #sh = wb['Sheet']
# #chart = LineChart()
#
# for n in data:
#     sh.append([n[0], n[1]])
#
# cdata = Reference(sh, min_col=2, min_row=1, max_row=sh.max_row)
# cat = Reference(sh, min_col=1, min_row=1, max_row=sh.max_row)
# chart = BarChart()
#
# chart.add_data(cdata)
# chart.set_categories(cat)
# sh.add_chart(chart, 'F4')
#
# wb.save('cctv.xlsx')

#5 (cctv증가율:  2013년도 이전/(2014년+2015년+2016년)
# cctv_increased = 0
# for n in data:
#     cctv_increased = n[2] / (n[3]+n[4]+n[5])
#     print(n[0], cctv_increased)

def cctv_inc():
    for dt in data:
        print(dt[0], round(dt[2] / (dt[3]+dt[4]+dt[5]), 2))

cctv_inc()


def saveDB():
    db = sqlite3.connect('cctv.db')
    db.execute('create table if not exists cctv(name varchar(20), stot int)')
    tdata = [(n[0], n[1]) for n in data]
    db.executemany('insert into cctv(name, stot) values(?,?)', tdata)

    db.commit()
    db.close()

saveDB()
