import openpyxl
from openpyxl.chart import BarChart, Reference, LineChart

data =[]
def inputData(sh):
    while True:
        name = input('이름:')
        kor = int( input('국어:') )
        eng = int( input('영어:') )
        math = int( input('수학:') )
        #data.append( (name, kor, eng,math) )
        sh.append([name, kor, eng, math])
        yn= input('계속 입력하시겠습니까(y/n)?')
        if yn=='n':
            break

def showHeader(sh):
    sh.append(('이름', '국어', '영어', '수학'))

def outputData():
    print("="*30)
    print('이름','국어','영어','수학',sep='\t')
    print("="*30)
    for n,k,e,m in data:
        print(  n,k,e,m,sep='\t\t' )

# entry point
wb = openpyxl.Workbook()
sh = wb.active

#wb = openpyxl.load_workbook('my1.xlsx')
#sh = wb['Sheet']

showHeader(sh)
inputData(sh)

chart = BarChart()
cdata = Reference(sh, min_col=2, max_col=4, min_row=1, max_row=sh.max_row)
cat = Reference(sh, min_col=1, min_row=2, max_row=sh.max_row)
chart.add_data(cdata, titles_from_data=True)
chart.set_categories(cat)
sh.add_chart(chart, 'F1')

for c1, c2, c3, c4 in sh:
    print(c1.value, c2.value, c3.value, c4.value)


wb.save('grade.xlsx')


#wb1 = openpyxl.load_workbook('grade.xlsx')
#sh1 = wb['Sheet']

wb.close()

#outputData()

